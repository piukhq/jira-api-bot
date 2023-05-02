import os

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from jira_enums import SPREADSHEET_BASE_DIR, Worksheets


def write_data_to_spreadsheet(squad_info: list, filename: str, worksheet: Worksheets):
    if not os.path.isdir(SPREADSHEET_BASE_DIR):
        os.mkdir(SPREADSHEET_BASE_DIR)

    file_path = os.path.join(SPREADSHEET_BASE_DIR, filename)
    try:
        wb = load_workbook(filename=file_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb["Sheet"])

    ws = wb.create_sheet(title=worksheet.value)

    data = [list(squad_info[0].keys())]
    for squad_dict in squad_info:
        data.append(list(squad_dict.values()))

    for row in data:
        ws.append(row)

    wb.save(file_path)
